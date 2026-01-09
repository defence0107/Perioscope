import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split, StratifiedKFold
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
from sklearn.metrics import (accuracy_score, roc_auc_score, precision_score,
                             recall_score, f1_score, confusion_matrix, roc_curve,
                             precision_recall_curve, average_precision_score, r2_score)
from sklearn.calibration import calibration_curve
import joblib
import os
import shap
from scipy.stats import bootstrap  # 用于计算置信区间
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# 设置中文字体，确保中文正常显示
plt.rcParams["font.family"] = ["Arial", "sans-serif"]
plt.rcParams["axes.unicode_minus"] = False  # 解决负号显示问题

# 文件路径
file_path = 'E:/D/linjun.xlsx'

# 文件存在性检查
if os.path.exists(file_path):
    print("文件存在")
    datasets = pd.read_excel(file_path, header=0)
else:
    print("文件不存在，请检查路径和文件名")
    exit()

# 数据预处理
Y = datasets.iloc[:, 0]
X = datasets.iloc[:, 1:]
feature_names = X.columns.tolist()

# 划分数据集：训练集70%，验证集15%，测试集15%
X_temp, X_Test, Y_temp, Y_Test = train_test_split(X, Y, test_size=0.15, random_state=0)
X_Train, X_Val, Y_Train, Y_Val = train_test_split(X_temp, Y_temp, test_size=(0.15 / 0.85), random_state=0)

# 特征标准化
sc_X = StandardScaler()
X_Train = pd.DataFrame(sc_X.fit_transform(X_Train), columns=feature_names)
X_Val = pd.DataFrame(sc_X.transform(X_Val), columns=feature_names)
X_Test = pd.DataFrame(sc_X.transform(X_Test), columns=feature_names)


def calculate_metrics(y_true, y_pred, y_proba, set_name=""):
    """计算并返回所有评估指标"""
    cm = confusion_matrix(y_true, y_pred)
    tn, fp, fn, tp = cm.ravel()

    metrics = {
        "Accuracy": accuracy_score(y_true, y_pred),
        "AUC": roc_auc_score(y_true, y_proba),
        "Precision": precision_score(y_true, y_pred),
        "Recall": recall_score(y_true, y_pred),
        "F1": f1_score(y_true, y_pred),
        "Specificity": tn / (tn + fp)
    }

    # 打印指标
    print(f"{set_name}指标:")
    for metric, value in metrics.items():
        print(f"  {metric}: {value:.4f}")

    return metrics


# 实验不同树数量的影响
n_estimators_list = range(1, 101, 1)

# 创建DataFrame保存所有指标
results_df = pd.DataFrame(columns=['n_estimators',
                                   'Train_Accuracy', 'Train_AUC', 'Train_Precision', 'Train_Recall', 'Train_F1',
                                   'Train_Specificity',
                                   'Val_Accuracy', 'Val_AUC', 'Val_Precision', 'Val_Recall', 'Val_F1',
                                   'Val_Specificity',
                                   'Test_Accuracy', 'Test_AUC', 'Test_Precision', 'Test_Recall', 'Test_F1',
                                   'Test_Specificity'])

for n_est in n_estimators_list:
    print(f"\n正在训练 n_estimators = {n_est}...")

    # 初始化分类器
    rf_classifier = RandomForestClassifier(
        n_estimators=n_est,
        criterion='entropy',
        random_state=0,
        max_depth=10,
        min_samples_leaf=10,
        n_jobs=-1,
        class_weight='balanced'
    )

    # 训练模型
    rf_classifier.fit(X_Train, Y_Train)

    # 训练集预测
    Y_Train_Pred = rf_classifier.predict(X_Train)
    y_train_proba = rf_classifier.predict_proba(X_Train)[:, 1]

    # 验证集预测
    Y_Val_Pred = rf_classifier.predict(X_Val)
    y_val_proba = rf_classifier.predict_proba(X_Val)[:, 1]

    # 测试集预测
    Y_Test_Pred = rf_classifier.predict(X_Test)
    y_test_proba = rf_classifier.predict_proba(X_Test)[:, 1]

    # 计算指标
    train_metrics = calculate_metrics(Y_Train, Y_Train_Pred, y_train_proba, "训练集")
    val_metrics = calculate_metrics(Y_Val, Y_Val_Pred, y_val_proba, "验证集")
    test_metrics = calculate_metrics(Y_Test, Y_Test_Pred, y_test_proba, "测试集")

    # 将结果添加到DataFrame
    results_df.loc[len(results_df)] = {
        'n_estimators': n_est,
        'Train_Accuracy': train_metrics['Accuracy'],
        'Train_AUC': train_metrics['AUC'],
        'Train_Precision': train_metrics['Precision'],
        'Train_Recall': train_metrics['Recall'],
        'Train_F1': train_metrics['F1'],
        'Train_Specificity': train_metrics['Specificity'],
        'Val_Accuracy': val_metrics['Accuracy'],
        'Val_AUC': val_metrics['AUC'],
        'Val_Precision': val_metrics['Precision'],
        'Val_Recall': val_metrics['Recall'],
        'Val_F1': val_metrics['F1'],
        'Val_Specificity': val_metrics['Specificity'],
        'Test_Accuracy': test_metrics['Accuracy'],
        'Test_AUC': test_metrics['AUC'],
        'Test_Precision': test_metrics['Precision'],
        'Test_Recall': test_metrics['Recall'],
        'Test_F1': test_metrics['F1'],
        'Test_Specificity': test_metrics['Specificity']
    }

    # 打印当前进度
    print(f"当前进度 ({n_est}/100)")
    print("-" * 50)

# 保存结果到Excel文件
output_file = "random_forest_metrics.xlsx"

# 如果文件已存在，先删除
if os.path.exists(output_file):
    os.remove(output_file)

# 创建新的Excel文件
wb = Workbook()
ws = wb.active

# 将DataFrame写入Excel
for r in dataframe_to_rows(results_df, index=False, header=True):
    ws.append(r)

# 保存Excel文件
wb.save(output_file)
print(f"\n所有指标已保存到 {output_file}")

# 可视化指标变化
plt.figure(figsize=(14, 8))
metrics_to_plot = ['Accuracy', 'AUC', 'F1']
colors = ['#1f77b4', '#ff7f0e', '#2ca02c']

for idx, metric in enumerate(metrics_to_plot):
    plt.plot(results_df['n_estimators'], results_df[f'Train_{metric}'],
             color=colors[idx], linestyle='--', label=f'Train {metric}')
    plt.plot(results_df['n_estimators'], results_df[f'Val_{metric}'],
             color=colors[idx], linestyle=':', label=f'Val {metric}')
    plt.plot(results_df['n_estimators'], results_df[f'Test_{metric}'],
             color=colors[idx], linewidth=2, label=f'Test {metric}')

plt.title('模型性能随树数量变化')
plt.xlabel('树的数量')
plt.ylabel('分数')
plt.xticks(n_estimators_list)
plt.grid(True, alpha=0.3)
plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()
plt.savefig('model_performance_plot.png')
plt.show()

# 保存最终模型（假设n_estimators=100为最佳模型）
final_model = rf_classifier
joblib.dump(final_model, 'random_forest_model_final.pkl', compress=True)


# ==================== 绘制混淆矩阵 ====================
import os
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix


def plot_confusion_matrix(y_true, y_pred, set_name="", output_dir="confusion_matrices"):
    """绘制并保存混淆矩阵，使用Arial字体"""
    os.makedirs(output_dir, exist_ok=True)

    # 设置全局字体为Arial
    plt.rcParams["font.family"] = ["Arial", "sans-serif"]

    cm = confusion_matrix(y_true, y_pred)

    plt.figure(figsize=(10, 8))
    # 获取heatmap的轴对象
    ax = sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                     xticklabels=['0', '1'],
                     yticklabels=['0', '1'],
                     annot_kws={"size": 25})

    # 设置标题和标签
    plt.title(f'Confusion Matrix({set_name})', fontsize=25)
    plt.xlabel('Predicted Label', fontsize=25)
    plt.ylabel('True Label', fontsize=25)

    # 使用轴对象的tick_params方法设置刻度
    ax.tick_params(axis='x', labelsize=25)  # x轴刻度
    ax.tick_params(axis='y', labelsize=22)  # y轴刻度

    plt.tight_layout()
    plt.savefig(f"{output_dir}/confusion_matrix_{set_name.lower()}.png", dpi=300)
    plt.close()


# 绘制所有混淆矩阵
plot_confusion_matrix(Y_Train, final_model.predict(X_Train), "Training")
plot_confusion_matrix(Y_Val, final_model.predict(X_Val), "Validation")
plot_confusion_matrix(Y_Test, final_model.predict(X_Test), "Test")

# ==================== 绘制ROC曲线 ====================
def plot_roc_curves(y_true_list, y_proba_list, set_names, output_path="roc_curves.png"):
    """
    绘制多组数据的ROC曲线

    Parameters:
    y_true_list (list): 真实标签列表（每个元素为numpy数组）
    y_proba_list (list): 预测概率列表（每个元素为numpy数组，正类概率）
    set_names (list): 数据集名称列表（如['Train', 'Val', 'Test']）
    output_path (str): 图像保存路径
    """
    plt.figure(figsize=(10, 8))

    # 定义颜色和样式
    colors = {
        'Train': {'color': '#88C4D7', 'marker': 'o', 'ls': '-'},
        'Validation': {'color': '#D0EAD5', 'marker': 's', 'ls': '--'},
        'Test': {'color': '#AFADD2', 'marker': '^', 'ls': ':'}
    }

    for y_true, y_proba, name in zip(y_true_list, y_proba_list, set_names):
        fpr, tpr, _ = roc_curve(y_true, y_proba)
        roc_auc = roc_auc_score(y_true, y_proba)

        # 使用指定的颜色和样式
        style = colors.get(name, {'color': 'gray', 'ls': '-'})
        plt.plot(fpr, tpr, lw=6, label=f'{name} (AUC = {roc_auc:.4f})',
                 color=style['color'], linestyle=style['ls'])
        plt.plot([0, 1], [0, 1], color='gray', lw=3, linestyle='--')  # 对角线

    plt.xlim([-0.05, 1.05])
    plt.ylim([-0.05, 1.05])
    # 增大坐标轴标签和刻度字体大小
    plt.xlabel('False Positive Rate (FPR)', fontsize=20, fontname='Arial')
    plt.ylabel('True Positive Rate (TPR)', fontsize=20, fontname='Arial')
    plt.xticks(fontsize=20)  # 增大x轴刻度
    plt.yticks(fontsize=20)  # 增大y轴刻度
    plt.title('ROC Curve', fontsize=20, fontname='Arial')

    # 调整图例位置，向上移动（通过bbox_to_anchor微调）
    # bbox_to_anchor的前两个值是相对坐标，(1, 0.1)表示右侧10%高度位置
    plt.legend(loc='lower right', bbox_to_anchor=(1, 0.005),
               fontsize=20, markerscale=1.5, frameon=False)

    # 添加灰色半透明网格线
    plt.grid(True, color='gray', alpha=0.3)

    # 去除上轴和右侧轴
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(output_path, dpi=600)
    plt.close()
    print(f"ROC curve saved to {output_path}")


# 生成ROC曲线数据
y_train_proba_final = final_model.predict_proba(X_Train)[:, 1]
y_val_proba_final = final_model.predict_proba(X_Val)[:, 1]
y_test_proba_final = final_model.predict_proba(X_Test)[:, 1]

# 准备数据
y_true_list = [Y_Train, Y_Val, Y_Test]
y_proba_list = [y_train_proba_final, y_val_proba_final, y_test_proba_final]
set_names = ['Train', 'Validation', 'Test']

# 调用函数绘制ROC曲线
plot_roc_curves(y_true_list, y_proba_list, set_names, output_path="roc_curves.png")


# ==================== 绘制PR曲线 ====================
def plot_pr_curves(y_true_list, y_proba_list, set_names, output_path="pr_curves.png"):
    """绘制多组数据的PR曲线"""
    plt.figure(figsize=(10, 8))

    # 定义颜色和样式
    colors = {
        'Train': {'color': '#88C4D7', 'marker': 'o', 'ls': '-'},
        'Validation': {'color': '#D0EAD5', 'marker': 's', 'ls': '--'},
        'Test': {'color': '#AFADD2', 'marker': '^', 'ls': ':'}
    }

    for y_true, y_proba, name in zip(y_true_list, y_proba_list, set_names):
        precision, recall, _ = precision_recall_curve(y_true, y_proba)
        avg_precision = average_precision_score(y_true, y_proba)

        # 使用指定的颜色和样式
        style = colors.get(name, {'color': 'gray', 'ls': '-'})
        plt.plot(recall, precision, lw=6,
                 label=f'{name} (AP = {avg_precision:.4f})',
                 color=style['color'], linestyle=style['ls'])

    plt.xlim([-0.05, 1.05])
    plt.ylim([-0.05, 1.05])
    # 增大坐标轴标签和刻度字体大小
    plt.xlabel('Recall', fontsize=20, fontname='Arial')
    plt.ylabel('Precision', fontsize=20, fontname='Arial')
    plt.xticks(fontsize=20)  # 增大x轴刻度
    plt.yticks(fontsize=20)  # 增大y轴刻度
    plt.title('Precision-Recall Curve', fontsize=20, fontname='Arial')
    # 增大图例符号并设置字体大小，取消边框
    plt.legend(loc='lower left', fontsize=20, markerscale=1.5, frameon=False)

    # 添加灰色半透明网格线
    plt.grid(True, color='gray', alpha=0.3)

    # 去除上轴和右侧轴
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(output_path, dpi=600)
    plt.close()
    print(f"PR curve saved to {output_path}")


# 调用函数绘制PR曲线
plot_pr_curves(y_true_list, y_proba_list, set_names, output_path="pr_curves.png")


# ==================== 绘制校准曲线 ====================
def plot_calibration_curves(y_true_list, y_proba_list, set_names, output_path="calibration_curves.png"):
    """绘制多组数据的校准曲线（无曲线符号和图例符号）"""
    plt.figure(figsize=(10, 8))

    # 定义颜色和样式（移除了marker参数）
    colors = {
        'Train': {'color': '#88C4D7', 'ls': '-'},
        'Validation': {'color': '#D0EAD5', 'ls': '--'},
        'Test': {'color': '#AFADD2', 'ls': ':'}
    }

    # 绘制理想校准线
    plt.plot([0, 1], [0, 1], linestyle='--', label='Perfectly calibrated',
             color='gray', lw=3)

    for y_true, y_proba, name in zip(y_true_list, y_proba_list, set_names):
        prob_true, prob_pred = calibration_curve(y_true, y_proba, n_bins=12)

        # 使用指定的颜色和样式（不设置marker）
        style = colors.get(name, {'color': 'gray', 'ls': '-'})
        plt.plot(prob_pred, prob_true, linestyle=style['ls'],  # 移除了marker参数
                 label=f'{name}', color=style['color'], lw=6)  # 移除了markersize

    # 增大坐标轴标签和刻度字体大小
    plt.xlabel('Mean predicted probability', fontsize=20, fontname='Arial')
    plt.ylabel('Fraction of positives', fontsize=20, fontname='Arial')
    plt.xticks(fontsize=20)  # 增大x轴刻度
    plt.yticks(fontsize=20)  # 增大y轴刻度
    plt.title('Calibration Curve', fontsize=20, fontname='Arial')
    # 增大图例字体大小，取消边框和标记缩放
    plt.legend(loc='upper left', fontsize=20, frameon=False)  # 移除了markerscale

    # 添加灰色半透明网格线
    plt.grid(True, color='gray', alpha=0.3)

    # 去除上轴和右侧轴
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(output_path, dpi=600)
    plt.close()
    print(f"Calibration curve saved to {output_path}")


# 调用函数绘制校准曲线
plot_calibration_curves(y_true_list, y_proba_list, set_names, output_path="calibration_curves.png")




# ==================== 预测概率分布图 ====================
def plot_probability_distributions(y_true_list, y_proba_list, set_names):
    """绘制正负类预测概率分布图"""
    for y_true, y_proba, name in zip(y_true_list, y_proba_list, set_names):
        plt.figure(figsize=(10, 6))

        # 分离正负类概率
        pos_proba = y_proba[y_true == 1]
        neg_proba = y_proba[y_true == 0]

        # 绘制直方图
        plt.hist(pos_proba, bins=20, alpha=0.5, color='red', label='正类')
        plt.hist(neg_proba, bins=20, alpha=0.5, color='blue', label='负类')

        plt.xlabel('预测概率', fontsize=12)
        plt.ylabel('样本数量', fontsize=12)
        plt.title(f'{name}预测概率分布', fontsize=14)
        plt.legend(loc='upper center', fontsize=10)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(f"probability_distribution_{name}.png", dpi=300)
        plt.close()
    print("预测概率分布图已保存")


# 调用函数绘制概率分布图
plot_probability_distributions(y_true_list, y_proba_list, set_names)

# ================ SHAP分析 ================
# 创建保存目录
shap_dir = "shap_analysis_results"
os.makedirs(shap_dir, exist_ok=True)

# 使用最终模型
final_model = rf_classifier

# 准备解释数据
X_explain = X_Test  # 形状应为 (样本数, 特征数)

# 验证数据维度
print("\n=== 数据验证 ===")
print(f"特征数量: {X_explain.shape[1]}")
print(f"样本数量: {X_explain.shape[0]}")

# 初始化SHAP解释器
explainer = shap.TreeExplainer(final_model)

# 计算SHAP值
shap_values = explainer.shap_values(X_explain)

# 调试输出
print("\n=== SHAP值结构 ===")
print(f"SHAP值类型: {type(shap_values)}")
print(f"SHAP值形状: {np.array(shap_values).shape}")

# 提取正确维度的SHAP值（二分类场景）
if isinstance(shap_values, list) and len(shap_values) == 2:
    print("\n检测到二分类模型（列表形式）")
    shap_values_positive = np.array(shap_values[1])  # 形状应为 (样本数, 特征数)
elif isinstance(shap_values, np.ndarray) and shap_values.ndim == 3 and shap_values.shape[2] == 2:
    print("\n检测到二分类模型（三维数组形式）")
    shap_values_positive = shap_values[:, :, 1]  # 提取正类的SHAP值
else:
    print("\n模型可能是多分类，需要调整索引")
    exit()

# 最终维度验证
try:
    assert shap_values_positive.shape == X_explain.shape
except AssertionError as e:
    print(f"\n!! 维度验证失败 !!")
    print(f"SHAP值形状: {shap_values_positive.shape}")
    print(f"数据形状: {X_explain.shape}")
    print("可能原因：模型实际输出维度与预期不符")
    exit()

# ================= 可视化 =================
# 转换数据格式
X_explain_array = X_explain.values  # 转换为numpy数组

# 设置字体为Arial
plt.rcParams["font.family"] = ["Arial", "sans-serif"]

# 1. 特征重要性图
plt.figure(figsize=(12, 6))
shap.summary_plot(shap_values_positive,
                  X_explain_array,
                  feature_names=feature_names,
                  plot_type="bar",
                  show=False)
plt.title("特征重要性 (SHAP值)", fontsize=14)
plt.tight_layout()
plt.savefig(f"{shap_dir}/1_feature_importance.png", dpi=300)
plt.close()

# 2. 特征效应散点图
plt.figure(figsize=(14, 8))
shap.summary_plot(shap_values_positive,
                  X_explain_array,
                  feature_names=feature_names,
                  show=False)
plt.title("特征效应图", fontsize=14)
plt.tight_layout()
plt.savefig(f"{shap_dir}/2_feature_effects.png", dpi=300)
plt.close()

# 3. 热力图
shap_explanation = shap.Explanation(values=shap_values_positive, feature_names=feature_names, data=X_explain_array)

plt.figure(figsize=(25, 12 + len(feature_names) * 0.3))
plt.rcParams.update({
    'font.size': 6,
    'axes.titlesize': 8,
    'axes.labelsize': 5,
    'xtick.labelsize': 5,
    'ytick.labelsize': 5,
})
shap.plots.heatmap(
    shap_explanation,
    max_display=len(feature_names),
    show=False
)
plt.title("SHAP值热力图")

# 手动调整热力图坐标轴
ax = plt.gca()
ax.set_xlabel(ax.get_xlabel(), fontsize=5)
ax.set_ylabel(ax.get_ylabel(), fontsize=5)
for tick in ax.get_xticklabels():
    tick.set_fontsize(5)
for tick in ax.get_yticklabels():
    tick.set_fontsize(4)  # 特征名称通常较长，使用更小的字体

plt.tight_layout()
plt.subplots_adjust(left=0.4)
plt.savefig(f"{shap_dir}/3_feature_heatmap.png", dpi=300)
plt.close()

# ==================== 2. 计算并导出所有特征的SHAP重要性及权重占比到Excel ====================
print("\n" + "=" * 80)
print("【SHAP特征重要性及权重占比（基于SHAP值绝对值均值）- 导出至Excel】")
print("=" * 80)

# 2.1 计算SHAP特征重要性（绝对值均值：消除正负向抵消，反映总影响力）
# 确保shap_values_positive已通过之前的SHAP分析代码计算（形状：[样本数, 特征数]）
try:
    # 按特征维度（axis=0）计算每个特征的SHAP值绝对值均值
    shap_importance = np.mean(np.abs(shap_values_positive), axis=0)
    feature_count = len(shap_importance)
    print(f"成功计算 {feature_count} 个特征的SHAP重要性\n")

    # 验证特征名称与重要性数量匹配
    if len(feature_names) != feature_count:
        raise ValueError(f"特征名称数量（{len(feature_names)}）与SHAP重要性数量（{feature_count}）不匹配，请检查特征列表！")
except NameError:
    print("错误：shap_values_positive未定义，请先执行SHAP分析代码！")
    exit()
except Exception as e:
    print(f"计算SHAP重要性时出错：{str(e)}")
    exit()

# 2.2 计算每个特征的权重占比（单个特征SHAP重要性 / 所有特征总SHAP重要性 * 100）
total_shap_importance = np.sum(shap_importance)
# 处理极端情况（总重要性为0，避免除以0错误）
if total_shap_importance == 0:
    shap_importance_ratio = np.zeros_like(shap_importance)
    print("警告：所有特征的SHAP重要性总和为0，权重占比将全部设为0\n")
else:
    shap_importance_ratio = (shap_importance / total_shap_importance) * 100  # 转为百分比

# 2.3 按SHAP重要性降序排序（确保所有特征都被包含，无遗漏）
sorted_shap_idx = np.argsort(shap_importance)[::-1]  # 降序索引（从大到小）
# 提取排序后的特征数据
sorted_data = {
    "排名": range(1, feature_count + 1),
    "特征名称": [feature_names[i] for i in sorted_shap_idx],
    "SHAP重要性（绝对值均值）": shap_importance[sorted_shap_idx],
    "权重占比(%)": shap_importance_ratio[sorted_shap_idx],
    "权重占比(格式化)": [f"{ratio:.2f}%" for ratio in shap_importance_ratio[sorted_shap_idx]],
    "累积权重占比(%)": np.cumsum(shap_importance_ratio[sorted_shap_idx]),  # 新增累积占比，便于识别关键特征
    "累积权重占比(格式化)": [f"{np.cumsum(shap_importance_ratio[sorted_shap_idx])[idx]:.2f}%"
                             for idx in range(feature_count)]
}

# 2.4 构建完整DataFrame（包含所有特征数据）
shap_ratio_df = pd.DataFrame(sorted_data)

# 2.5 导出到Excel（优化格式：添加数据说明、调整列宽、冻结表头）
# 生成带时间戳的文件名，避免覆盖
output_excel = f"shap_feature_importance_ratio_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # 工作表1：特征权重占比数据（主表）
    shap_ratio_df.to_excel(writer, sheet_name="特征权重占比明细", index=False)
    # 工作表2：数据说明（提升可读性）
    info_data = {
        "说明项": [
            "数据来源",
            "SHAP重要性计算方式",
            "权重占比计算逻辑",
            "累积权重占比含义",
            "排序规则",
            "数据生成时间",
            "特征总数"
        ],
        "详细说明": [
            "基于随机森林模型的SHAP分析结果（测试集数据）",
            "每个特征SHAP值的绝对值在所有样本上的均值（消除正负向抵消）",
            "单个特征SHAP重要性 / 所有特征SHAP重要性总和 × 100%",
            "按重要性排序后，前N个特征的权重占比之和（用于识别关键特征）",
            "按SHAP重要性降序排列（从影响最大到最小）",
            pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(feature_count) + " 个"
        ]
    }
    info_df = pd.DataFrame(info_data)
    info_df.to_excel(writer, sheet_name="数据说明", index=False)

    # 优化Excel格式（调整列宽、冻结表头）
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # 处理"特征权重占比明细"工作表
    ws1 = writer.sheets["特征权重占比明细"]
    # 冻结表头（第一行）
    ws1.freeze_panes = "A2"
    # 设置表头样式（加粗、居中）
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    for col in range(1, len(shap_ratio_df.columns) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_alignment
        # 调整列宽（根据列名长度和数据类型适配）
        col_name = shap_ratio_df.columns[col - 1]
        if "特征名称" in col_name:
            ws1.column_dimensions[get_column_letter(col)].width = 30
        elif "占比" in col_name and "格式化" in col_name:
            ws1.column_dimensions[get_column_letter(col)].width = 15
        elif "SHAP重要性" in col_name or "占比" in col_name:
            ws1.column_dimensions[get_column_letter(col)].width = 20
        else:
            ws1.column_dimensions[get_column_letter(col)].width = 10

    # 处理"数据说明"工作表
    ws2 = writer.sheets["数据说明"]
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 80
    # 表头样式
    for col in range(1, 3):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_alignment

# 2.6 打印导出结果与数据摘要
print("📊 数据导出摘要：")
print(f"   - 导出文件：{output_excel}")
print(f"   - 包含特征数：{feature_count} 个（无遗漏）")
print(f"   - Excel包含工作表：「特征权重占比明细」「数据说明」")
print(f"   - 明细字段：排名、特征名称、SHAP重要性、权重占比、累积权重占比（含格式化版本）")
print("\n✅ SHAP特征权重占比数据已成功导出到Excel！")

# 可选：打印前10个重要特征的预览（快速查看关键信息）
print("\n" + "-" * 50)
print("前10个重要特征预览（完整数据见Excel）：")
print("-" * 50)
preview_df = shap_ratio_df[["排名", "特征名称", "权重占比(格式化)", "累积权重占比(格式化)"]].head(10)
for _, row in preview_df.iterrows():
    print(
        f"排名{row['排名']:2d} | 特征：{row['特征名称']:<20} | 权重占比：{row['权重占比(格式化)']:<8} | 累积占比：{row['累积权重占比(格式化)']}")

print(f"\nSHAP分析成功！结果保存在 {shap_dir}/")

# ==================== 训练集和测试集预测概率的线性回归图 ====================
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import r2_score

plt.figure(figsize=(10, 8))

# 获取预测概率（分类模型使用predict_proba获取正类概率）
y_train_proba = final_model.predict_proba(X_Train)[:, 1]  # 正类概率
y_test_proba = final_model.predict_proba(X_Test)[:, 1]    # 正类概率

# 定义颜色和样式
train_color = '#B7DBE3'  # 训练集颜色
test_color = '#C4C3DE'   # 测试集颜色
trendline_colors = {
    'Train': '#2B7AB8',  # 训练集趋势线颜色（更深）
    'Test': '#6A5ACD'    # 测试集趋势线颜色（更深）
}

# 绘制训练集预测概率散点
plt.scatter(Y_Train, y_train_proba,
            c=train_color, alpha=0.6, s=50,
            edgecolor='white', linewidth=0.5,
            label='Training Set')

# 绘制测试集预测概率散点
plt.scatter(Y_Test, y_test_proba,
            c=test_color, alpha=0.6, s=50,
            edgecolor='white', linewidth=0.5,
            label='Test Set')

# 添加训练集概率趋势线
sns.regplot(x=Y_Train, y=y_train_proba,
            scatter=False,
            line_kws={
                'color': trendline_colors['Train'],
                'linestyle': '-',
                'linewidth': 2.5,
                'alpha': 0.8
            })

# 添加测试集概率趋势线
sns.regplot(x=Y_Test, y=y_test_proba,
            scatter=False,
            line_kws={
                'color': trendline_colors['Test'],
                'linestyle': '-',
                'linewidth': 2.5,
                'alpha': 0.8
            })

# 添加理想校准线（预测概率与真实类别完全匹配）
plt.plot([0, 1], [0, 1], 'k--', linewidth=1.5, label='Perfect Calibration')

# 计算R²分数（评估预测概率与真实值的拟合程度）
train_r2 = r2_score(Y_Train, y_train_proba)
test_r2 = r2_score(Y_Test, y_test_proba)

# 创建自定义图例，包含R²分数
legend_elements = [
    plt.Line2D([0], [0], marker='o', color='w',
               markerfacecolor=train_color, markersize=10,
               label=f'Training (R²={train_r2:.3f})'),
    plt.Line2D([0], [0], marker='o', color='w',
               markerfacecolor=test_color, markersize=10,
               label=f'Test (R²={test_r2:.3f})'),
    plt.Line2D([0], [0], color=trendline_colors['Train'],
               linewidth=2.5, label='Training Trend'),
    plt.Line2D([0], [0], color=trendline_colors['Test'],
               linewidth=2.5, label='Test Trend'),
    plt.Line2D([0], [0], color='k', linestyle='--',
               linewidth=1.5, label='Ideal Line')
]

plt.legend(handles=legend_elements, loc='upper left',
           frameon=True, framealpha=0.9)

# 设置坐标轴标签和标题
plt.xlabel('True Class Labels (0/1)', fontsize=12, fontweight='bold')
plt.ylabel('Predicted Probability (Class 1)', fontsize=12, fontweight='bold')
plt.title('Predicted Probability vs True Class Comparison',
          fontsize=14, fontweight='bold')

# 调整坐标轴范围和刻度
plt.xlim(-0.05, 1.05)
plt.ylim(-0.05, 1.05)
plt.xticks([0, 0.5, 1])
plt.yticks([0, 0.5, 1])
plt.grid(True, alpha=0.2, linestyle='--')

# 美化图形：去除顶部和右侧边框
ax = plt.gca()
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

# 保存图形
plt.tight_layout()
plt.savefig('probability_calibration_plot.png', dpi=300, bbox_inches='tight')
plt.close()
print("预测概率校准图已保存为 probability_calibration_plot.png")


# 新增：计算指标及其置信区间的函数
def calculate_metrics_with_ci(model, X, y, dataset_name, metrics_functions, n_bootstraps=1000, ci_level=0.95):
    """
    计算指标点估计值和置信区间

    参数:
    model: 训练好的模型
    X: 特征数据 (DataFrame)
    y: 真实标签 (Series)
    dataset_name: 数据集名称 (str)
    metrics_functions: 指标计算函数字典 {指标名: 函数}
    n_bootstraps: 自助抽样次数
    ci_level: 置信水平

    返回:
    包含点估计和置信区间的DataFrame
    """
    results = []
    n_samples = len(y)
    rng = np.random.default_rng(42)  # 固定随机种子确保可重复性

    # 定义自助抽样函数
    def bootstrap_sample():
        indices = rng.choice(n_samples, size=n_samples, replace=True)
        return X.iloc[indices], y.iloc[indices]

    # 计算原始数据点估计
    y_proba = model.predict_proba(X)[:, 1]
    y_pred = model.predict(X)
    point_estimates = {name: func(y, y_pred, y_proba) for name, func in metrics_functions.items()}

    # 自助抽样计算置信区间
    for metric_name, metric_func in metrics_functions.items():
        bootstrap_values = []
        for _ in range(n_bootstraps):
            X_boot, y_boot = bootstrap_sample()
            y_proba_boot = model.predict_proba(X_boot)[:, 1]
            y_pred_boot = model.predict(X_boot)
            bootstrap_values.append(metric_func(y_boot, y_pred_boot, y_proba_boot))

        # 计算置信区间
        sorted_values = np.sort(bootstrap_values)
        lower = sorted_values[int((1 - ci_level) / 2 * n_bootstraps)]
        upper = sorted_values[int((1 + ci_level) / 2 * n_bootstraps)]

        results.append({
            'Metric': metric_name,
            'Point Estimate': point_estimates[metric_name],
            'CI Lower': lower,
            'CI Upper': upper,
            'Dataset': dataset_name  # 使用显式传入的数据集名称
        })

    return pd.DataFrame(results)


# 定义需要计算的指标及其计算函数
metrics_functions = {
    'Accuracy': lambda y, y_pred, y_proba: accuracy_score(y, y_pred),
    'AUC': lambda y, y_pred, y_proba: roc_auc_score(y, y_proba),
    'Precision': lambda y, y_pred, y_proba: precision_score(y, y_pred),
    'Recall': lambda y, y_pred, y_proba: recall_score(y, y_pred),
    'F1': lambda y, y_pred, y_proba: f1_score(y, y_pred),
    'Specificity': lambda y, y_pred, y_proba:
    confusion_matrix(y, y_pred).ravel()[0] / (
            confusion_matrix(y, y_pred).ravel()[0] + confusion_matrix(y, y_pred).ravel()[1])
}

# 计算各数据集的置信区间
# 直接传入数据集名称，而不是依赖DataFrame的name属性
ci_train = calculate_metrics_with_ci(final_model, X_Train, Y_Train, "Train", metrics_functions)
ci_val = calculate_metrics_with_ci(final_model, X_Val, Y_Val, "Validation", metrics_functions)
ci_test = calculate_metrics_with_ci(final_model, X_Test, Y_Test, "Test", metrics_functions)

# 合并结果
all_ci_results = pd.concat([ci_train, ci_val, ci_test], ignore_index=True)

# 写入单独的Excel文件
ci_output_path = 'metrics_confidence_intervals.xlsx'
with pd.ExcelWriter(ci_output_path, engine='openpyxl') as writer:
    # 创建透视表，按指标和数据集组织结果
    pivot_table = all_ci_results.pivot_table(
        index='Metric',
        columns='Dataset',
        values=['Point Estimate', 'CI Lower', 'CI Upper']
    )

    # 确保列按有意义的顺序排列
    ordered_datasets = ['Train', 'Validation', 'Test']
    ordered_columns = [(metric_type, dataset)
                       for metric_type in ['Point Estimate', 'CI Lower', 'CI Upper']
                       for dataset in ordered_datasets]

    # 重新排序列
    pivot_table = pivot_table[ordered_columns]

    # 写入Excel
    pivot_table.to_excel(writer, sheet_name='CI Results', float_format='%.4f')

    # 添加单独的工作表，按数据集分开
    for dataset in ['Train', 'Validation', 'Test']:
        dataset_results = all_ci_results[all_ci_results['Dataset'] == dataset]
        dataset_results = dataset_results[['Metric', 'Point Estimate', 'CI Lower', 'CI Upper']]
        dataset_results.to_excel(writer, sheet_name=f'CI_{dataset}', index=False, float_format='%.4f')

print(f"\n置信区间计算完成，结果已保存至 {ci_output_path}")

# 可选：打印结果摘要
print("\n=== 置信区间摘要 ===")
for dataset in ['Train', 'Validation', 'Test']:
    subset = all_ci_results[all_ci_results['Dataset'] == dataset]
    print(f"\n{dataset} 数据集:")
    for _, row in subset.iterrows():
        print(f"{row['Metric']}: {row['Point Estimate']:.4f} ({row['CI Lower']:.4f}-{row['CI Upper']:.4f})")


def calculate_metrics_with_cv(model, X, y, metrics_functions, n_splits=5, random_state=42):
    """
    执行五折交叉验证并计算指标的点估计和置信区间
    """
    # 初始化分层K折交叉验证
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    fold_results = []

    # 遍历每个折叠
    for fold, (train_idx, val_idx) in enumerate(skf.split(X, y), 1):
        # 划分训练/验证集
        X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
        y_train, y_val = y.iloc[train_idx], y.iloc[val_idx]

        # 折叠内标准化（避免数据泄漏）
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_val_scaled = scaler.transform(X_val)

        # 训练模型
        model.fit(X_train_scaled, y_train)

        # 预测
        y_proba = model.predict_proba(X_val_scaled)[:, 1]
        y_pred = model.predict(X_val_scaled)

        # 计算指标
        fold_metrics = {
            "Fold": fold,
            **{name: func(y_val, y_pred, y_proba) for name, func in metrics_functions.items()}
        }
        fold_results.append(fold_metrics)

    # 转换为DataFrame
    fold_df = pd.DataFrame(fold_results).set_index("Fold")

    # 计算点估计（均值）和置信区间（基于交叉验证结果的自助法）
    ci_results = []
    for metric in metrics_functions.keys():
        values = fold_df[metric].values
        point_estimate = np.mean(values)

        # 修正：使用正确的参数名statistic和data
        boot_result = bootstrap(
            data=(values,),  # 数据参数名应为data
            statistic=lambda x: np.mean(x),  # 统计函数参数名应为statistic
            n_resamples=1000,
            random_state=random_state
        )
        ci_lower, ci_upper = boot_result.confidence_interval

        ci_results.append({
            "Metric": metric,
            "Point Estimate": point_estimate,
            "CI Lower": ci_lower,
            "CI Upper": ci_upper,
            "Dataset": "5-Fold CV"
        })

    return pd.DataFrame(ci_results), fold_df


# ==================== 执行五折交叉验证 ====================
# 初始化模型（使用最终确定的超参数，如n_estimators=100）
cv_model = RandomForestClassifier(
    n_estimators=100,
    criterion='entropy',
    random_state=0,
    max_depth=10,
    min_samples_leaf=10,
    n_jobs=-1,
    class_weight='balanced'
)

# 执行交叉验证并获取结果
cv_ci, cv_folds = calculate_metrics_with_cv(
    model=cv_model,
    X=X_Train,
    y=Y_Train,
    metrics_functions=metrics_functions,
    n_splits=5,
    random_state=0
)

# ==================== 保存交叉验证结果 ====================
# 创建输出文件
cv_output_file = "cv_metrics_with_ci.xlsx"
with pd.ExcelWriter(cv_output_file, engine='openpyxl') as writer:
    # 写入置信区间
    cv_ci.to_excel(writer, sheet_name="CI Summary", float_format="%.4f")

    # 写入各折叠详细结果
    cv_folds.to_excel(writer, sheet_name="Fold Results", float_format="%.4f")

print(f"\n五折交叉验证完成，结果保存至 {cv_output_file}")
print("\n各折叠指标详情:")
print(cv_folds)
print("\n置信区间摘要:")
print(cv_ci[["Metric", "Point Estimate", "CI Lower", "CI Upper"]])

